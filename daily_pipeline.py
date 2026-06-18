#!/usr/bin/env python3
"""
Daily Job Scrub Pipeline - mechanical tasks.

Implements Tasks 0, 1, 2, the Task 3 triage, and Task 4 from
daily-job-scrub-pipeline.md as a script, so Cowork doesn't have to
re-derive this dedup/filter/triage logic (and re-read every posting)
each run.

Two phases, run with a company-vetting + fit-assessment pass in between:

    python3 daily_pipeline.py phase1
        Tasks 0, 1, 2, Task 3 triage.
        Writes: seen-postings.xlsx, company-tracker.xlsx (PENDING rows),
                job-title-filters.xlsx (Unclassified Titles log)
        Outputs: handoff_companies.json - companies needing full vetting
                 this run (<= max_new_companies_per_run), for Cowork to
                 run company-vetting-subagent.md against.

    [Cowork runs company-vetting-subagent.md for each company in
     handoff_companies.json, updating company-tracker.xlsx]

    python3 daily_pipeline.py phase2
        Task 4.
        Writes: job-listings.xlsx ("Shortlist")
        Outputs: handoff_fit_assessments.json - Shortlist rows needing
                 assessment this run (<= max_fit_assessments_per_run),
                 for Cowork to run job-fit-assessment-subagent.md against.

    [Cowork runs job-fit-assessment-subagent.md for each row in
     handoff_fit_assessments.json, writing to job-listings.xlsx ("Fit Assessment")]

    python3 daily_pipeline.py build-review
        Syncs any Status/Skip Reason edits B made in "Shortlist" back
        to "All Listings" (the permanent record), then rebuilds "Shortlist"
        from scratch: Shortlist rows not yet Applied/Skipped, joined with
        Fit Assessment, ranked best-first, capped at review_list_size.
        Prints the final run summary (Initial Retrieval / new companies to
        review / final review list, per README "After the Run").

Both phases print a one-line-per-task progress summary to stdout.
"""

import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

import greenhouse_api as gh

DATA_DIR = Path(__file__).resolve().parent

_BOARD_ROTATION_FILE = "board_rotation_state.json"


def _load_board_rotation_offset():
    """Returns the index into the Active-board list to START from this run,
    persisted across runs in board_rotation_state.json. Without this, every
    run processes the same boards in company-boards.xlsx row order until
    max_postings_per_run is hit - on a 16-board list where rows 1-6 alone
    can exceed the cap, boards further down (often the carefully
    department-filtered PURSUE companies, since WATCH companies with no
    filter tend to get added/listed first) never get attempted at all,
    every single run. Starting from a rotating offset means every board
    gets a turn across enough runs, instead of the same handful winning
    forever by virtue of row position."""
    path = DATA_DIR / _BOARD_ROTATION_FILE
    if not path.exists():
        return 0
    try:
        return json.loads(path.read_text()).get("next_offset", 0)
    except (json.JSONDecodeError, OSError):
        return 0


def _save_board_rotation_offset(offset, total_boards):
    if total_boards <= 0:
        return
    path = DATA_DIR / _BOARD_ROTATION_FILE
    path.write_text(json.dumps({"next_offset": offset % total_boards}))

TODAY = date.today()

def _validate_today():
    """Sanity-check the system clock. If the date looks implausible (before
    2026 or more than a week in the future vs. a hardcoded floor), Cowork
    should ask B for the current date rather than stamping records wrong."""
    floor = date(2026, 1, 1)
    if TODAY < floor:
        print(f"WARNING: System clock returned {TODAY}, which is before {floor}.")
        print("The sandbox clock may be wrong. Set the correct date by running:")
        print('  PIPELINE_DATE="2026-06-15" python3 daily_pipeline.py phase1')
        raise SystemExit(1)

if os.environ.get("PIPELINE_DATE"):
    try:
        TODAY = date.fromisoformat(os.environ["PIPELINE_DATE"])
    except ValueError:
        print(f"ERROR: PIPELINE_DATE={os.environ['PIPELINE_DATE']!r} is not a valid YYYY-MM-DD date.")
        raise SystemExit(1)
else:
    _validate_today()


# ---------------------------------------------------------------------------
# Config (parsed from the YAML block at the top of settings.md)
# ---------------------------------------------------------------------------

def load_config():
    text = (DATA_DIR / "settings.md").read_text()
    m = re.search(r"```yaml\n(.*?)\n```", text, re.DOTALL)
    if not m:
        raise RuntimeError("Could not find YAML config block in settings.md")
    cfg = yaml.safe_load(m.group(1))
    required = [
        "max_postings_per_run", "max_new_companies_per_run",
        "max_fit_assessments_per_run", "max_shortlist_per_company", "review_list_size",
        "max_posting_age_days",
        "shortlist_company_fit_floor",
        "comp_floor", "comp_target", "comp_unusually_high", "hourly_to_annual",
    ]
    missing = [k for k in required if k not in cfg]
    if missing:
        raise RuntimeError(f"settings.md config block missing keys: {missing}")
    return cfg


# ---------------------------------------------------------------------------
# Company Name Normalization (settings.md -> "Company Name Normalization")
# ---------------------------------------------------------------------------

_LEGAL_SUFFIXES = [
    ", Inc.", ", Inc", " Inc.", " Inc",
    ", LLC", " LLC",
    ", Ltd.", ", Ltd", " Ltd.", " Ltd",
    ", Corp.", ", Corp", " Corp.", " Corp", " Corporation",
    ", Co.", " Co.",
]


def normalize_company_name(name):
    if name is None:
        return ""
    n = str(name).strip()
    # Strip suffixes; repeat in case of "Foo, Inc., LLC" style double suffixes
    changed = True
    while changed:
        changed = False
        for suf in _LEGAL_SUFFIXES:
            if n.endswith(suf):
                n = n[: -len(suf)].rstrip().rstrip(",").strip()
                changed = True
    return n


def names_match(a, b):
    return normalize_company_name(a).casefold() == normalize_company_name(b).casefold()


# ---------------------------------------------------------------------------
# Title classification (job-title-filters.xlsx -> "Title Keywords")
#
# Evaluation order (see daily-job-scrub-pipeline.md Task 1):
#   1. SKIP keywords    - any match -> discard entirely
#   2. PURSUE keywords  - any match (no SKIP matched) -> Role Match = PURSUE
#   3. CHECK keywords   - any match (no SKIP/PURSUE matched) -> Role Match = CHECK
#   4. no match at all  -> Role Match = CHECK, log to Unclassified Titles
# ---------------------------------------------------------------------------

def load_title_keywords(wb):
    ws = wb["Title Keywords"]
    skip, pursue, check = [], [], []
    for r in range(2, ws.max_row + 1):
        kw = ws.cell(row=r, column=1).value
        cls = ws.cell(row=r, column=2).value
        if not kw or not cls:
            continue
        kw_lower = str(kw).strip().lower()
        cls = str(cls).strip().upper()
        if cls == "SKIP":
            skip.append(kw_lower)
        elif cls == "PURSUE":
            pursue.append(kw_lower)
        elif cls == "CHECK":
            check.append(kw_lower)
    return skip, pursue, check


def classify_title(title, skip_kw, pursue_kw, check_kw):
    """Returns (role_match_or_None, reason). role_match is None if SKIPped."""
    t = (title or "").lower()
    for kw in skip_kw:
        if kw in t:
            return None, f"SKIP keyword: {kw}"
    for kw in pursue_kw:
        if kw in t:
            return "PURSUE", f"PURSUE keyword: {kw}"
    for kw in check_kw:
        if kw in t:
            return "CHECK", f"CHECK keyword: {kw}"
    return "CHECK", "no keyword match - unclassified"


# ---------------------------------------------------------------------------
# Location filter (location-filters.xlsx -> "Location Keywords")
# ---------------------------------------------------------------------------

def load_location_keywords(wb):
    ws = wb["Location Keywords"]
    keywords = []
    for r in range(2, ws.max_row + 1):
        kw = ws.cell(row=r, column=1).value
        if kw:
            keywords.append(str(kw).strip().lower())
    return keywords


# International terms that unambiguously mean "not in the US target area".
# Pure Python substring check — zero tokens, runs before the allowlist.
# Keep lowercase. These short-circuit location_passes immediately.
_INTL_BLOCKLIST = [
    # Countries
    ", canada", "canada", ", india", "india", ", uk", ", u.k.",
    "united kingdom", "england", "scotland", "wales",
    "ireland", "germany", "france", "netherlands", "spain", "italy",
    "poland", "sweden", "norway", "denmark", "finland", "switzerland",
    "austria", "belgium", "portugal", "czech", "romania", "hungary",
    "australia", "new zealand", "singapore", "japan", "china", "korea",
    "brazil", "mexico", "israel", "turkey", "ukraine", "russia",
    # Cities that are unambiguous (won't false-positive against US cities)
    "london", "berlin", "dublin", "warsaw", "reykjavik", "reykjavík",
    "amsterdam", "paris", "munich", "zurich", "stockholm", "oslo",
    "copenhagen", "helsinki", "brussels", "vienna", "madrid", "barcelona",
    "rome", "milan", "toronto", "vancouver", "montreal", "sydney",
    "melbourne", "tokyo", "seoul", "beijing", "shanghai", "tel aviv",
    "cape town", "nairobi", "bogotá", "ljubljana", "tallinn", "vilnius",
    # Region codes and suffixes Greenhouse uses
    "emea", "apac", "latam", "- uk", "- eu", "- europe",
    "remote, india", "remote, canada", "remote canada", "remote india",
    "remote uk", "remote, uk",
]


def location_passes(location_type, location_str, location_keywords):
    """Returns (passes: bool, matched_keyword: str|None).

    Three-stage check — fastest to slowest:
    1. Remote → pass immediately (None keyword = no novel-location logging needed).
    2. International blocklist → fail immediately. Pure Python, zero tokens.
    3. US allowlist → pass if any keyword from location-filters.xlsx matches.

    matched_keyword is set when the allowlist matched but the full raw string
    hasn't been explicitly catalogued — callers log that to Unrecognized Locations.
    """
    if location_type == "Remote":
        # Still check the international blocklist — "Remote, India" / "Remote, Canada"
        # should not pass just because they have "remote" in them.
        loc_check = (location_str or "").lower()
        for bad in _INTL_BLOCKLIST:
            if bad in loc_check:
                return False, None
        return True, None

    loc = (location_str or "").lower()

    # Stage 2: international blocklist — short-circuit before allowlist
    for bad in _INTL_BLOCKLIST:
        if bad in loc:
            return False, None

    # Stage 3: US allowlist (location-filters.xlsx → Location Keywords)
    for kw in location_keywords:
        if kw in loc:
            return True, kw

    return False, None


# ---------------------------------------------------------------------------
# Location Type / Employment Type mapping from raw board text
# (daily-job-scrub-pipeline.md Task 1 field-capture rules)
# ---------------------------------------------------------------------------

def map_location_type(raw_location):
    s = (raw_location or "").lower()
    if not s:
        return "Not Disclosed"

    # "Remote" must be the primary/sole location, not incidental in a
    # multi-office string like "SF • NYC • Remote" (which is really On-Site
    # with a maybe-remote option). Greenhouse multi-location strings often
    # use • / | ; separators — if "remote" appears AND there are other
    # location segments, treat as Hybrid, not Remote.
    if "remote" in s:
        separators = ["•", "|", ";", "/"]
        is_multi = any(sep in s for sep in separators)
        # Also flag "United States" as a standalone Remote indicator
        if not is_multi:
            return "Remote"
        # e.g. "Remote - US" or "Remote US" with no other segments = Remote
        stripped = s.replace("remote", "").replace("-", "").replace("us", "").strip()
        if not stripped or stripped in ("", "–", "—", "·", " "):
            return "Remote"
        return "Hybrid"

    if "hybrid" in s:
        return "Hybrid"

    # A specific city/state string with no remote/hybrid → On-Site.
    if "," in s or any(ch.isalpha() for ch in s):
        return "On-Site"
    return "Not Disclosed"


def map_employment_type(raw_text):
    """raw_text: any free text available (title, metadata) that might hint
    at employment type. Greenhouse's list endpoint (content=false) doesn't
    reliably expose this, so this is best-effort; default Not Disclosed.
    Short abbreviations (FT/PT) use word boundaries - "ft" as a bare
    substring matches inside "software", "draft", etc."""
    s = (raw_text or "").lower()
    if any(k in s for k in ["full time", "full-time", "fulltime", "permanent"]):
        return "Full-time"
    if re.search(r"\bft\b", s):
        return "Full-time"
    if any(k in s for k in ["contractor", "contract", "1099", "c2c"]):
        return "Contract"
    if any(k in s for k in ["part time", "part-time", "parttime"]):
        return "Part-time"
    if re.search(r"\bpt\b", s):
        return "Part-time"
    if "freelance" in s:
        return "Freelance"
    if "temporary" in s or "temp" in s:
        return "Temporary"
    return "Not Disclosed"


# ---------------------------------------------------------------------------
# Job ID extraction (for fallback / non-API sources)
# ---------------------------------------------------------------------------

def extract_job_id(url):
    if not url:
        return None
    m = re.search(r"gh_jid=(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/jobs/(\d+)(?:[/?]|$)", url)
    if m:
        return m.group(1)
    return None


# ---------------------------------------------------------------------------
# XLSX helpers
# ---------------------------------------------------------------------------

DATE_COLUMNS = {
    "job-listings.xlsx":    {"All Listings": ["Date Found", "Posted Date"],
                             "Shortlist":    ["Date Found"]},
    "seen-postings.xlsx":   {"Seen Postings": ["First Seen Date", "Last Seen Date", "Applied Date"]},
    "company-tracker.xlsx": {"Company Tracker": ["Date Checked", "Next Review Date"]},
    "job-title-filters.xlsx": {"Unclassified Titles": ["Date Found"]},
    "location-filters.xlsx":  {"Unrecognized Locations": ["Date Found"]},
}


def migrate_dates_in_workbook(wb, sheet_col_map):
    """One-time / idempotent: convert any datetime cells in the given
    (sheet -> [col_header]) map to plain date objects. Strings that look
    like ISO dates are also converted. Leaves non-date values untouched.
    Returns total cells changed."""
    changed = 0
    for sheet_name, col_headers in sheet_col_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        hmap = header_map(ws)
        for col_header in col_headers:
            col = hmap.get(col_header)
            if col is None:
                continue
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                v = cell.value
                if isinstance(v, datetime):
                    cell.value = v.date()
                    cell.number_format = "YYYY-MM-DD"
                    changed += 1
                elif isinstance(v, str) and re.match(r"\d{4}-\d{2}-\d{2}", v):
                    try:
                        cell.value = date.fromisoformat(v[:10])
                        cell.number_format = "YYYY-MM-DD"
                        changed += 1
                    except ValueError:
                        pass
    return changed


def header_map(ws):
    """{header_text: 1-based column index}"""
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}


def rows_as_dicts(ws):
    """Yield (row_index, {header: value}) for each data row.
    datetime values are normalized to date — openpyxl always loads date
    serial numbers as datetime regardless of cell format, since Excel stores
    dates and datetimes identically at the file level."""
    hmap = header_map(ws)
    for r in range(2, ws.max_row + 1):
        d = {}
        for h, c in hmap.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime) and v.hour == 0 and v.minute == 0 and v.second == 0:
                v = v.date()
            d[h] = v
        if all(v is None for v in d.values()):
            continue
        yield r, d


def _write_cell(ws, row, col, value):
    """Write value to cell. For date objects, also set number_format so
    openpyxl/Excel stores and displays them as YYYY-MM-DD, not datetime."""
    cell = ws.cell(row=row, column=col, value=value)
    if isinstance(value, date) and not isinstance(value, datetime):
        cell.number_format = "YYYY-MM-DD"


def append_dict_row(ws, rowdict):
    hmap = header_map(ws)
    r = ws.max_row + 1
    # Guard against the conditional-formatting max_row inflation bug:
    # find the true next empty row by checking column A.
    while ws.cell(row=r - 1, column=1).value is None and r > 2:
        r -= 1
    for h, c in hmap.items():
        _write_cell(ws, r, c, rowdict.get(h))
    return r


def set_dict_row(ws, row_index, rowdict):
    hmap = header_map(ws)
    for h, c in hmap.items():
        if h in rowdict:
            _write_cell(ws, row_index, c, rowdict[h])


# ---------------------------------------------------------------------------
# Task 0: Sync Applied Status
# ---------------------------------------------------------------------------

def sync_review_to_shortlist(wb_listings):
    """Read job-listings.xlsx -> "Shortlist" (the human-facing ranked list —
    B marks Status=Applied/Skipped + Skip Reason there). Write those values
    back to the matching "All Listings" row (by Job ID, falling back to URL) —
    All Listings is the permanent record; "Shortlist" gets rebuilt fresh
    every run by build_review_sheet(). Returns count of All Listings rows updated."""
    if "Shortlist" not in wb_listings.sheetnames:
        return 0
    ws_review = wb_listings["Shortlist"]
    ws_short = wb_listings["All Listings"]

    by_jid, by_url = {}, {}
    for r, d in rows_as_dicts(ws_short):
        if d.get("Job ID"):
            by_jid[str(d["Job ID"])] = r
        if d.get("URL"):
            by_url[d["URL"]] = r

    updated = 0
    for _, rd in rows_as_dicts(ws_review):
        status = str(rd.get("Status") or "").strip()
        if status.lower() not in ("applied", "skipped"):
            continue
        jid, url = rd.get("Job ID"), rd.get("URL")
        target = by_jid.get(str(jid)) if jid else None
        if target is None and url:
            target = by_url.get(url)
        if target is None:
            continue
        # Normalize to the canonical capitalization regardless of what B typed.
        canonical = "Applied" if status.lower() == "applied" else "Skipped"
        set_dict_row(ws_short, target, {"Status": canonical, "Skip Reason": rd.get("Skip Reason")})
        updated += 1
    return updated


def task0_sync_applied_status(wb_listings, wb_seen):
    """For Shortlist rows marked Status = Applied, set Applied?=Yes / Applied
    Date=today on the matching seen-postings row (matched by Job ID, then
    Posting URL). Returns count of seen-postings rows updated."""
    ws_short = wb_listings["All Listings"]
    ws_seen = wb_seen["Seen Postings"]

    # Build lookup: job_id -> row_index, url -> row_index
    by_jid, by_url = {}, {}
    for r, d in rows_as_dicts(ws_seen):
        jid = d.get("Job ID")
        if jid:
            by_jid[str(jid)] = r
        url = d.get("Posting URL")
        if url:
            by_url[url] = r

    updated = 0
    for _, srow in rows_as_dicts(ws_short):
        if str(srow.get("Status")).strip().lower() != "applied":
            continue
        jid = srow.get("Job ID")
        url = srow.get("URL")
        target_row = by_jid.get(str(jid)) if jid else None
        if target_row is None and url:
            target_row = by_url.get(url)
        if target_row is None:
            continue
        seen_d = {h: ws_seen.cell(row=target_row, column=c).value for h, c in header_map(ws_seen).items()}
        changed = False
        if str(seen_d.get("Applied?")).strip().lower() != "yes":
            set_dict_row(ws_seen, target_row, {"Applied?": "Yes"})
            changed = True
        if not seen_d.get("Applied Date"):
            set_dict_row(ws_seen, target_row, {"Applied Date": TODAY})
            changed = True
        if changed:
            updated += 1
    return updated


# ---------------------------------------------------------------------------
# Task 1: Pull, Dedup & Title-Filter New Listings
# ---------------------------------------------------------------------------
# Greenhouse-specific fetch/parse logic now lives in greenhouse_api.py
# (gh.fetch_jobs, gh.process_job, etc.) so a second ATS platform can be
# added later without touching this file's pipeline logic.



def task1_pull_dedup_filter(cfg, wb_boards, wb_sources, wb_seen, wb_title_filters, wb_loc_filters,
                             wb_tracker=None, fetch_fn=None):
    if fetch_fn is None:
        fetch_fn = gh.fetch_jobs
    ws_src = wb_sources["Source Sites"]
    enabled_sources = {
        str(d.get("Source Name")).strip()
        for _, d in rows_as_dicts(ws_src)
        if str(d.get("Status")).strip().lower() == "enabled"
    }

    ws_boards = wb_boards["Company Boards"]
    boards = [
        (d.get("Company Name"), d.get("Source"), d.get("Board URL"),
         [x.strip() for x in str(d.get("Departments") or "").split(",") if x.strip()])
        for _, d in rows_as_dicts(ws_boards)
        if str(d.get("Status")).strip().lower() == "active" and str(d.get("Source")).strip() in enabled_sources
    ]

    # Tier boards so PURSUE companies get priority over WATCH/unvetted ones,
    # rather than processing strictly in company-boards.xlsx row order. A
    # handful of unfiltered WATCH boards (no Departments configured, so they
    # pull every job on the board) can otherwise consume the entire
    # max_postings_per_run cap before a single PURSUE company - the ones
    # actually worth pursuing - gets touched, every single run, regardless
    # of where the rotation offset currently points.
    if wb_tracker is not None:
        ws_tracker = wb_tracker["Company Tracker"]
        pursue_companies = {
            normalize_company_name(d.get("Company Name")).casefold()
            for _, d in rows_as_dicts(ws_tracker)
            if str(d.get("Status")).strip().upper() == "PURSUE"
        }
        tier1 = [b for b in boards if normalize_company_name(b[0]).casefold() in pursue_companies]
        tier2 = [b for b in boards if normalize_company_name(b[0]).casefold() not in pursue_companies]
    else:
        tier1, tier2 = boards, []

    ws_seen = wb_seen["Seen Postings"]
    ws_needslink = wb_seen["Needs Link"]
    seen_hmap = header_map(ws_seen)

    by_jid, by_url, by_company_title = {}, {}, {}
    for r, d in rows_as_dicts(ws_seen):
        if d.get("Job ID"):
            by_jid[str(d["Job ID"])] = r
        if d.get("Posting URL"):
            by_url[d["Posting URL"]] = r
        key = (normalize_company_name(d.get("Company Name")).casefold(), str(d.get("Job Title") or "").strip().lower())
        by_company_title.setdefault(key, []).append(r)

    skip_kw, pursue_kw, check_kw = load_title_keywords(wb_title_filters)
    ws_unclassified = wb_title_filters["Unclassified Titles"]
    existing_unclassified = {str(d.get("Job Title") or "").strip().lower() for _, d in rows_as_dicts(ws_unclassified)}

    loc_kw = load_location_keywords(wb_loc_filters)
    ws_unrecognized_locs = wb_loc_filters["Unrecognized Locations"]
    existing_unrecognized_locs = {str(d.get("Location") or "").strip().lower() for _, d in rows_as_dicts(ws_unrecognized_locs)}

    max_pull = cfg["max_postings_per_run"]

    # Rotate TIER 1 (PURSUE boards) so they take turns going first across
    # runs, rather than always starting at row 1 within that tier. Tier 2
    # (WATCH/unvetted boards) is appended after, unrotated - it only gets
    # processed with whatever cap budget tier 1 doesn't use, so fairness
    # within tier 2 matters much less than guaranteeing tier 1 always gets
    # first crack at the cap. Wraps around at the end of tier 1's list.
    total_tier1 = len(tier1)
    offset = _load_board_rotation_offset() % total_tier1 if total_tier1 else 0
    tier1 = tier1[offset:] + tier1[:offset]
    boards = tier1 + tier2
    total_boards = len(boards)

    counts = {
        "boards_attempted": 0, "board_errors": [], "fetched": 0,
        "no_url": 0, "already_seen": 0, "new": 0, "location_excluded": 0,
        "stale_excluded": 0,
        "title_skip": 0, "survivors": 0, "capped": False,
        "unclassified_titles": 0, "unrecognized_locations": 0,
        "fetch_sources": {},  # company_name -> "direct"/"cache"/"none"/"mixed/cache"
    }
    survivors = []
    pulled_total = 0
    boards_processed = 0

    for company_name, source, board_url, dept_ids in boards:
        if pulled_total >= max_pull:
            counts["capped"] = True
            break
        slug = gh.slug_from_board_url(board_url)
        counts["boards_attempted"] += 1
        boards_processed += 1
        result = fetch_fn(slug, dept_ids=dept_ids) if dept_ids else fetch_fn(slug)
        # fetch_fn may be gh.fetch_jobs (returns 3-tuple with source) or a
        # test double (returns 2-tuple) - handle both so existing tests
        # that pass a custom fetch_fn don't need to change.
        if len(result) == 3:
            jobs, err, fetch_source = result
        else:
            jobs, err = result
            fetch_source = "unknown"
        counts["fetch_sources"][company_name] = fetch_source
        if err:
            counts["board_errors"].append(f"{company_name} ({slug}): {err}")
            continue

        for job in jobs:
            if pulled_total >= max_pull:
                counts["capped"] = True
                break
            rec = gh.process_job(job, company_name, source, map_location_type, normalize_company_name, map_employment_type)

            pulled_total += 1
            counts["fetched"] += 1

            if not rec["URL"]:
                counts["no_url"] += 1
                append_dict_row(ws_needslink, {
                    "Date Found": TODAY, "Company Name": rec["Company Name"],
                    "Job Title": rec["Job Title"], "Source": rec["Source"], "Notes": "",
                })
                continue

            jid = rec["Job ID"]
            existing_row = by_jid.get(jid) if jid else None
            if existing_row is None:
                existing_row = by_url.get(rec["URL"])

            if existing_row:
                counts["already_seen"] += 1
                times_seen = ws_seen.cell(row=existing_row, column=seen_hmap["Times Seen"]).value or 0
                set_dict_row(ws_seen, existing_row, {
                    "Last Seen Date": TODAY,
                    "Times Seen": times_seen + 1,
                })
                continue

            counts["new"] += 1
            key = (rec["Company Name"].casefold(), rec["Job Title"].strip().lower())
            notes = ""
            for prev_row in by_company_title.get(key, []):
                if str(ws_seen.cell(row=prev_row, column=seen_hmap["Applied?"]).value).strip().lower() == "yes":
                    notes = "possible repost after application"
                    break

            new_seen_row = {
                "Posting URL": rec["URL"], "Job ID": jid, "Company Name": rec["Company Name"],
                "Job Title": rec["Job Title"], "Department": rec["Department"], "Source": rec["Source"],
                "First Seen Date": TODAY, "Last Seen Date": TODAY,
                "Times Seen": 1, "Applied?": "No", "Notes": notes,
            }

            passes, matched_kw = location_passes(rec["Location Type"], rec["Location"], loc_kw)
            if not passes:
                counts["location_excluded"] += 1
                excl_note = f"Excluded - location ({rec['Location']})"
                new_seen_row["Notes"] = f"{notes}; {excl_note}" if notes else excl_note
                new_r = append_dict_row(ws_seen, new_seen_row)
                if jid:
                    by_jid[jid] = new_r
                by_url[rec["URL"]] = new_r
                continue

            # Hard age cutoff: postings older than max_posting_age_days never
            # become survivors, full stop - not a ranking penalty, an actual
            # exclusion. Without this, a 2-month-old posting could still get
            # selected for fit assessment (if fresher competition was thin)
            # and recommended on the Shortlist, which is exactly what
            # happened with several Figma postings in a real run - the only
            # staleness handling that existed before this was a mild ranking
            # penalty in the fit-assessment selection, applied too late and
            # too gently to prevent that. Missing/unparseable Posted Date is
            # NOT excluded here (better to show an undated posting than
            # silently drop real new listings due to a parsing gap) - that
            # case is instead logged to Needs Link-style visibility via the
            # Notes field below, so it's not silently treated as either
            # fresh or stale.
            posted_date_obj = rec.get("_posted_date_obj")
            max_age = cfg.get("max_posting_age_days")
            if max_age is not None and posted_date_obj is not None:
                age_days = (TODAY - posted_date_obj).days
                if age_days > max_age:
                    counts["stale_excluded"] += 1
                    excl_note = f"Excluded - stale ({age_days}d old, posted {posted_date_obj.isoformat()})"
                    new_seen_row["Notes"] = f"{notes}; {excl_note}" if notes else excl_note
                    new_r = append_dict_row(ws_seen, new_seen_row)
                    if jid:
                        by_jid[jid] = new_r
                    by_url[rec["URL"]] = new_r
                    continue

            # Log novel passing location strings (passed via keyword substring
            # but the full raw string hasn't been explicitly catalogued).
            if matched_kw is not None:  # None = Remote, no logging needed
                loc_key = (rec["Location"] or "").strip().lower()
                if loc_key and loc_key not in existing_unrecognized_locs and loc_key != matched_kw:
                    append_dict_row(ws_unrecognized_locs, {
                        "Date Found": TODAY, "Location": rec["Location"],
                        "Company Name": rec["Company Name"], "Job Title": rec["Job Title"],
                        "Reviewed?": "No", "Notes": f"passed via keyword: {matched_kw}",
                    })
                    existing_unrecognized_locs.add(loc_key)
                    counts["unrecognized_locations"] += 1

            role_match, reason = classify_title(rec["Job Title"], skip_kw, pursue_kw, check_kw)
            if role_match is None:
                counts["title_skip"] += 1
                new_r = append_dict_row(ws_seen, new_seen_row)
                if jid:
                    by_jid[jid] = new_r
                by_url[rec["URL"]] = new_r
                continue

            if "no keyword match" in reason:
                title_key = rec["Job Title"].strip().lower()
                if title_key not in existing_unclassified:
                    append_dict_row(ws_unclassified, {
                        "Date Found": TODAY, "Job Title": rec["Job Title"],
                        "Company Name": rec["Company Name"], "Reviewed?": "No", "Notes": "",
                    })
                    existing_unclassified.add(title_key)
                    counts["unclassified_titles"] += 1

            new_r = append_dict_row(ws_seen, new_seen_row)
            if jid:
                by_jid[jid] = new_r
            by_url[rec["URL"]] = new_r

            counts["survivors"] += 1
            survivor = {k: v for k, v in rec.items() if not k.startswith("_")}
            survivor["Role Match"] = role_match
            survivors.append(survivor)

    # Only advance the offset by however many TIER 1 boards were consumed -
    # if the run had enough cap budget to spill into tier 2, that doesn't
    # mean tier 1 should skip further ahead than it actually progressed.
    tier1_consumed = min(boards_processed, total_tier1)
    _save_board_rotation_offset(offset + tier1_consumed, total_tier1)
    counts["rotation_offset_this_run"] = offset
    counts["rotation_offset_next_run"] = (offset + tier1_consumed) % total_tier1 if total_tier1 else 0
    return survivors, counts


# ---------------------------------------------------------------------------
# Task 2: Group by Company
# ---------------------------------------------------------------------------

def task2_group_by_company(survivors, wb_tracker):
    """Returns a priority-ordered list of (company_name, listing_count).
    Companies with Status=PENDING in company-tracker are prepended
    (treated as highest priority) even if they have 0 new listings today."""
    from collections import Counter
    counts = Counter(s["Company Name"] for s in survivors)

    ws = wb_tracker["Company Tracker"]
    pending = []
    seen_norm = {normalize_company_name(n).casefold() for n in counts}
    for _, d in rows_as_dicts(ws):
        if str(d.get("Status") or "").strip().upper() == "PENDING":
            name = normalize_company_name(d.get("Company Name"))
            if name.casefold() not in seen_norm:
                pending.append(name)
                seen_norm.add(name.casefold())

    groups = [(name, cnt) for name, cnt in counts.items()]
    pending_set = {p.casefold() for p in pending}
    for p in pending:
        groups.append((p, 0))

    # Pending-from-prior-run first, then by listing count descending.
    groups.sort(key=lambda x: (0 if x[0].casefold() in pending_set else 1, -x[1]))
    return groups


# ---------------------------------------------------------------------------
# Task 3: Triage Companies
# ---------------------------------------------------------------------------

def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        try:
            return datetime.fromisoformat(v.strip()).date()
        except ValueError:
            return None
    return None


def task3_triage(cfg, company_groups, wb_tracker):
    """Single pass over company-tracker.xlsx. Returns:
        handoff: list of {"company": str, "listing_count": int} needing the
                 vetting subagent this run (within cap)
        resolved: list of {"company": str, "status": str} that resolved
                 instantly (no subagent call)
        pending_written: list of company names written/updated to PENDING
                 this run because they're over the cap
    """
    from openpyxl.utils import get_column_letter
    ws = wb_tracker["Company Tracker"]
    hmap = header_map(ws)

    tracker_row = {}
    for r, d in rows_as_dicts(ws):
        tracker_row[normalize_company_name(d.get("Company Name")).casefold()] = (r, d)

    resolved, needs_vetting = [], []
    for company, count in company_groups:
        key = normalize_company_name(company).casefold()
        entry = tracker_row.get(key)
        if entry is None:
            needs_vetting.append((company, count, None))
            continue
        row, d = entry
        status = str(d.get("Status") or "").strip().upper()
        if status == "BLACKLIST":
            resolved.append({"company": company, "status": "BLACKLIST"})
            continue
        if status in ("PURSUE", "WATCH"):
            date_checked = _parse_date(d.get("Date Checked"))
            interval = d.get("Review Interval (months)") or 6
            next_review = (date_checked + relativedelta(months=int(interval))) if date_checked else None
            if next_review and TODAY < next_review:
                resolved.append({"company": company, "status": status})
                continue
            needs_vetting.append((company, count, row))
            continue
        # PENDING or any other/unknown status -> needs vetting
        needs_vetting.append((company, count, row))

    cap = cfg["max_new_companies_per_run"]
    to_vet = needs_vetting[:cap]
    overflow = needs_vetting[cap:]

    pending_written = []
    for company, _, row in overflow:
        norm_name = normalize_company_name(company)
        review_interval = 6
        if row is not None:
            d = tracker_row[norm_name.casefold()][1]
            review_interval = d.get("Review Interval (months)") or 6
            r = row
        else:
            r = append_dict_row(ws, {"Company Name": norm_name})
        date_col = get_column_letter(hmap["Date Checked"])
        int_col = get_column_letter(hmap["Review Interval (months)"])
        set_dict_row(ws, r, {
            "Company Name": norm_name,
            "Status": "PENDING",
            "Company Fit": None,
            "Ghost Verdict": None,
            "Reason / Summary": "Awaiting vetting - capped this run",
            "Date Checked": TODAY,
            "Review Interval (months)": review_interval,
            "Next Review Date": f"=EDATE({date_col}{r},{int_col}{r})",
        })
        pending_written.append(norm_name)

    handoff = [{"company": normalize_company_name(c), "listing_count": cnt} for c, cnt, _ in to_vet]
    return handoff, resolved, pending_written


# ---------------------------------------------------------------------------
# Task 4: Filter Listings & Update Shortlist
# ---------------------------------------------------------------------------

_FIT_ORDER = ["Strong Match", "Solid Contender", "Long Shot", "Hard No"]


def company_fit_passes(company_fit, floor):
    if company_fit not in _FIT_ORDER or floor not in _FIT_ORDER:
        return False
    return _FIT_ORDER.index(company_fit) <= _FIT_ORDER.index(floor)


def _board_slug_for_company(wb_boards, company_name):
    ws = wb_boards["Company Boards"]
    for _, d in rows_as_dicts(ws):
        if names_match(d.get("Company Name"), company_name):
            return gh.slug_from_board_url(d.get("Board URL"))
    return None


def strip_html(html):
    """Strip HTML tags and decode common entities to plain text.
    Collapses whitespace. Returns at most 2000 chars — enough to skim
    requirements without blowing up a cell."""
    if not html:
        return ""
    import re as _re
    # Common entities
    text = html.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">") \
               .replace("&nbsp;", " ").replace("&#39;", "'").replace("&quot;", '"') \
               .replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n") \
               .replace("</p>", "\n").replace("</li>", "\n").replace("</h1>", "\n") \
               .replace("</h2>", "\n").replace("</h3>", "\n")
    text = _re.sub(r"<[^>]+>", "", text)
    text = _re.sub(r"[ \t]+", " ", text)
    text = _re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()[:2000]


def _shortlist_row_from_rec(rec, role_match, company_fit, description=""):
    return {
        "Date Found": TODAY, "Job Title": rec["Job Title"], "Department": rec.get("Department", "Not Disclosed"),
        "Role Match": role_match, "Company Name": rec["Company Name"], "Source": rec["Source"],
        "Location": rec["Location"], "Location Type": rec["Location Type"], "Pay Range": rec["Pay Range"],
        "Employment Type": rec["Employment Type"], "Posted Date": rec["Posted Date"],
        "URL": rec["URL"], "Job ID": rec["Job ID"], "Company Fit": company_fit,
        "Notes": "", "Job Description": description,
        "Status": "", "Skip Reason": "",
    }


def write_job_description(wb_listings, job_id, url, description_html):
    """Write stripped plain-text job description to the matching All Listings
    row (by Job ID, falling back to URL). Called from job-fit-assessment-subagent
    Step 5 after the full posting is fetched — that's when we have content=true.
    Returns True if a row was found and updated."""
    ws = wb_listings["All Listings"]
    hmap = header_map(ws)
    jid_col = hmap.get("Job ID")
    url_col = hmap.get("URL")
    desc_col = hmap.get("Job Description")
    if not desc_col:
        return False
    plain = strip_html(description_html)
    for r in range(2, ws.max_row + 1):
        match = False
        if job_id and jid_col:
            if str(ws.cell(row=r, column=jid_col).value) == str(job_id):
                match = True
        if not match and url and url_col:
            if ws.cell(row=r, column=url_col).value == url:
                match = True
        if match:
            ws.cell(row=r, column=desc_col, value=plain)
            return True
    return False


def task4_filter_shortlist(cfg, survivors, wb_tracker, wb_listings, wb_seen, wb_boards,
                            fetch_detail_fn=None):
    if fetch_detail_fn is None:
        fetch_detail_fn = gh.fetch_job_detail
    from collections import defaultdict

    ws_tracker = wb_tracker["Company Tracker"]
    tracker = {normalize_company_name(d.get("Company Name")).casefold(): d
               for _, d in rows_as_dicts(ws_tracker)}

    ws_short = wb_listings["All Listings"]
    existing_jids = {str(d.get("Job ID")) for _, d in rows_as_dicts(ws_short) if d.get("Job ID")}
    existing_urls = {d.get("URL") for _, d in rows_as_dicts(ws_short) if d.get("URL")}

    ws_seen = wb_seen["Seen Postings"]
    seen_hmap = header_map(ws_seen)

    floor = cfg["shortlist_company_fit_floor"]
    counts = {
        "candidates": 0, "added": 0, "discarded_blacklist": 0, "discarded_watch": 0,
        "pending_deferred": 0, "fit_floor_excluded": 0, "app_limit_capped": [],
        "revisit_added": 0, "revisit_cleared": 0, "revisit_refetch_failed": 0,
    }

    by_company = defaultdict(list)
    for s in survivors:
        by_company[normalize_company_name(s["Company Name"]).casefold()].append(s)

    for company_key, postings in by_company.items():
        d = tracker.get(company_key)
        status = str(d.get("Status") or "").strip().upper() if d else None
        company_name = d.get("Company Name") if d else normalize_company_name(postings[0]["Company Name"])

        if status == "BLACKLIST":
            counts["discarded_blacklist"] += len(postings)
            continue
        if status == "WATCH":
            counts["discarded_watch"] += len(postings)
            continue
        if status != "PURSUE":  # PENDING, or no tracker row at all (shouldn't happen post-triage, but safe)
            counts["pending_deferred"] += len(postings)
            for s in postings:
                _mark_seen_note(ws_seen, seen_hmap, s, "company pending - revisit")
            continue

        company_fit = d.get("Company Fit")
        if not company_fit_passes(company_fit, floor):
            counts["fit_floor_excluded"] += len(postings)
            continue

        counts["candidates"] += len(postings)
        candidates = sorted(postings, key=lambda s: 0 if s["Role Match"] == "PURSUE" else 1)

        app_limit = d.get("Application Limit")
        if app_limit not in (None, "", "[UNKNOWN]"):
            try:
                n = int(app_limit)
                if len(candidates) > n:
                    counts["app_limit_capped"].append((company_name, len(candidates), n))
                candidates = candidates[:n]
            except (ValueError, TypeError):
                pass

        for s in candidates:
            jid = s.get("Job ID")
            if (jid and str(jid) in existing_jids) or s["URL"] in existing_urls:
                continue
            append_dict_row(ws_short, _shortlist_row_from_rec(s, s["Role Match"], company_fit))
            if jid:
                existing_jids.add(str(jid))
            existing_urls.add(s["URL"])
            counts["added"] += 1

    _task4_revisit_pending(cfg, tracker, wb_seen, wb_listings, wb_boards, floor,
                            existing_jids, existing_urls, counts, fetch_detail_fn)
    return counts


def _mark_seen_note(ws_seen, seen_hmap, posting, note):
    """Find the seen-postings row for this posting (by Job ID, falling back
    to URL) and append `note` to its Notes if not already present."""
    jid_col, url_col, notes_col = seen_hmap["Job ID"], seen_hmap["Posting URL"], seen_hmap["Notes"]
    for r in range(2, ws_seen.max_row + 1):
        if posting.get("Job ID") and str(ws_seen.cell(row=r, column=jid_col).value) == str(posting["Job ID"]):
            match = True
        elif ws_seen.cell(row=r, column=url_col).value == posting.get("URL"):
            match = True
        else:
            match = False
        if match:
            existing = ws_seen.cell(row=r, column=notes_col).value or ""
            if note not in existing:
                new_note = f"{existing}; {note}" if existing else note
                ws_seen.cell(row=r, column=notes_col, value=new_note)
            return


def _task4_revisit_pending(cfg, tracker, wb_seen, wb_listings, wb_boards, floor,
                            existing_jids, existing_urls, counts, fetch_detail_fn):
    """Rescan seen-postings for rows noted "company pending - revisit" whose
    company has since resolved out of PENDING (this run's vetting may have
    just done that). Re-fetch full posting details (Location/Pay/etc. aren't
    stored in seen-postings) via the Greenhouse API using the stored Job ID."""
    ws_seen = wb_seen["Seen Postings"]
    seen_hmap = header_map(ws_seen)
    ws_short = wb_listings["All Listings"]

    for r, sd in rows_as_dicts(ws_seen):
        notes = str(sd.get("Notes") or "")
        if "company pending - revisit" not in notes:
            continue
        company_key = normalize_company_name(sd.get("Company Name")).casefold()
        d = tracker.get(company_key)
        if d is None:
            continue
        status = str(d.get("Status") or "").strip().upper()
        if status == "PENDING":
            continue  # still pending, leave the note as-is

        # Company resolved this run - clear the note either way
        cleared_notes = notes.replace("; company pending - revisit", "").replace("company pending - revisit", "").strip("; ").strip()

        if status in ("BLACKLIST", "WATCH"):
            ws_seen.cell(row=r, column=seen_hmap["Notes"], value=cleared_notes)
            counts["revisit_cleared"] += 1
            continue

        # status == PURSUE
        company_fit = d.get("Company Fit")
        if not company_fit_passes(company_fit, floor):
            ws_seen.cell(row=r, column=seen_hmap["Notes"], value=cleared_notes)
            counts["revisit_cleared"] += 1
            continue

        jid = sd.get("Job ID")
        if (jid and str(jid) in existing_jids) or sd.get("Posting URL") in existing_urls:
            ws_seen.cell(row=r, column=seen_hmap["Notes"], value=cleared_notes)
            counts["revisit_cleared"] += 1
            continue

        slug = _board_slug_for_company(wb_boards, sd.get("Company Name"))
        if slug and jid:
            result = fetch_detail_fn(slug, jid)
            if len(result) == 3:
                detail, err, _detail_source = result
            else:
                detail, err = result
        else:
            detail, err = None, "no board slug or Job ID available for refetch"

        if err or not detail:
            # Don't leave "pending - revisit" on this row forever once the
            # company has resolved - one refetch attempt, then hand off to
            # a manual check rather than retrying every future run.
            counts["revisit_refetch_failed"] += 1
            manual_note = "company resolved but posting could not be refreshed - check manually"
            new_notes = f"{cleared_notes}; {manual_note}" if cleared_notes else manual_note
            ws_seen.cell(row=r, column=seen_hmap["Notes"], value=new_notes)
            continue

        rec = gh.process_job(detail, sd.get("Company Name"), sd.get("Source"), map_location_type, normalize_company_name, map_employment_type)
        # Role Match isn't re-derived here - the title was already classified
        # PURSUE or CHECK when this posting first appeared (that's why it's
        # in seen-postings with a real Job Title at all; SKIPped titles never
        # get the "pending - revisit" note in the first place). Default to
        # CHECK, the more conservative of the two, since the original
        # classification isn't persisted in seen-postings.
        role_match = "CHECK"
        append_dict_row(ws_short, _shortlist_row_from_rec(rec, role_match, company_fit))
        if jid:
            existing_jids.add(str(jid))
        existing_urls.add(rec["URL"])
        ws_seen.cell(row=r, column=seen_hmap["Notes"], value=cleared_notes)
        counts["revisit_added"] += 1


# ---------------------------------------------------------------------------
# extract-cache CLI command
# ---------------------------------------------------------------------------

def cmd_extract_cache_batch(argv):
    """python3 daily_pipeline.py extract-cache-batch <spec1> <spec2> ... [--cache-dir DIR]

    Batch version of extract-cache: processes many board/dept/posting
    fetch results in ONE call instead of one tool invocation per file. A
    department-filtered board with N departments needs N separate cache
    files - Gitlab alone has 6 configured, Reddit 11, MongoDB 13 - so
    running extract-cache once per file multiplies tool-call overhead for
    no benefit; this collapses that into a single call.

    Each spec is colon-separated (not space-separated, since slugs/paths
    won't contain colons):
      board:<slug>:<input_file>
      dept:<slug>:<dept_id>:<input_file>
      posting:<job_id>:<input_file>

    Example:
      python3 daily_pipeline.py extract-cache-batch \\
        dept:gitlab:4115236002:gitlab_ai.txt \\
        dept:gitlab:4135580002:gitlab_arch.txt \\
        board:cloudflare:cloudflare_jobs.txt

    Prints one summary line per spec, plus a final count, rather than N
    separate "Wrote X job(s) to ..." outputs to scroll through."""
    if len(argv) < 1:
        print(cmd_extract_cache_batch.__doc__)
        sys.exit(1)
    override = argv[argv.index("--cache-dir") + 1] if "--cache-dir" in argv else None
    cdir = gh.cache_dir(override)
    specs = [a for a in argv if a != "--cache-dir" and a != override]

    results = []
    for spec in specs:
        parts = spec.split(":")
        kind = parts[0] if parts else ""
        try:
            if kind == "board":
                if len(parts) != 3:
                    results.append(f"SKIPPED (bad board spec, expected board:slug:file): {spec!r}")
                    continue
                _, slug, input_path = parts
                if not Path(input_path).exists():
                    results.append(f"ERROR ({slug}): file not found: {input_path}")
                    continue
                jobs, truncated = gh.extract_jobs_from_raw(Path(input_path).read_text(errors="replace"))
                out_dir = cdir / "boards"
                out_dir.mkdir(parents=True, exist_ok=True)
                out_path = out_dir / f"{slug}.json"
                out_path.write_text(json.dumps({"jobs": jobs}, indent=2))
                msg = f"board {slug}: {len(jobs)} job(s)"
                if truncated:
                    msg += " (truncated)"
                results.append(msg)

            elif kind == "dept":
                if len(parts) != 4:
                    results.append(f"SKIPPED (bad dept spec, expected dept:slug:dept_id:file): {spec!r}")
                    continue
                _, slug, dept_id, input_path = parts
                if not Path(input_path).exists():
                    results.append(f"ERROR ({slug} dept {dept_id}): file not found: {input_path}")
                    continue
                jobs, err = gh.extract_dept_from_raw(Path(input_path).read_text(errors="replace"))
                out_dir = cdir / "boards"
                out_dir.mkdir(parents=True, exist_ok=True)
                out_path = out_dir / f"{slug}-dept-{dept_id}.json"
                out_path.write_text(json.dumps({"jobs": jobs}, indent=2))
                msg = f"dept {slug}/{dept_id}: {len(jobs)} job(s)"
                if err:
                    msg += f" ({err})"
                results.append(msg)

            elif kind == "posting":
                if len(parts) != 3:
                    results.append(f"SKIPPED (bad posting spec, expected posting:job_id:file): {spec!r}")
                    continue
                _, job_id, input_path = parts
                if not Path(input_path).exists():
                    results.append(f"ERROR (posting {job_id}): file not found: {input_path}")
                    continue
                obj, status = gh.extract_posting_from_raw(Path(input_path).read_text(errors="replace"))
                out_dir = cdir / "postings"
                out_dir.mkdir(parents=True, exist_ok=True)
                out_path = out_dir / f"{job_id}.json"
                if status == "ok":
                    out_path.write_text(json.dumps(obj, indent=2))
                    results.append(f"posting {job_id}: ok")
                else:
                    out_path.write_text(json.dumps({"_status": status}))
                    results.append(f"posting {job_id}: {status}")
            else:
                results.append(f"SKIPPED (unknown kind {kind!r}, expected board/dept/posting): {spec!r}")
        except Exception as e:
            results.append(f"ERROR processing {spec!r}: {type(e).__name__}: {e}")

    print(f"Processed {len(specs)} spec(s):")
    for r in results:
        print(f"  {r}")
    error_count = sum(1 for r in results if r.startswith(("ERROR", "SKIPPED")))
    if error_count:
        print(f"\n{error_count} of {len(specs)} had issues - see above.")


def cmd_extract_cache(argv):
    """python3 daily_pipeline.py extract-cache board <slug> <input_file> [--cache-dir DIR]
       python3 daily_pipeline.py extract-cache posting <job_id> <input_file> [--cache-dir DIR]
       python3 daily_pipeline.py extract-cache dept <slug> <dept_id> <input_file> [--cache-dir DIR]

    Normalizes a raw web_fetch result (possibly header-prefixed, possibly
    truncated at ~94KB, possibly empty) into a clean JSON cache file the
    script can read. Run this once per board (or per department) before
    `phase1`, and once per posting before a `phase2` revisit rescan."""
    if len(argv) < 3 or argv[0] not in ("board", "posting", "dept"):
        print(cmd_extract_cache.__doc__)
        sys.exit(1)
    override = argv[argv.index("--cache-dir") + 1] if "--cache-dir" in argv else None
    cdir = gh.cache_dir(override)

    if argv[0] == "board":
        slug, input_path = argv[1], argv[2]
        jobs, truncated = gh.extract_jobs_from_raw(Path(input_path).read_text(errors="replace"))
        out_dir = cdir / "boards"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{slug}.json"
        out_path.write_text(json.dumps({"jobs": jobs}, indent=2))
        msg = f"Wrote {len(jobs)} job(s) to {out_path}"
        if truncated:
            msg += " (truncated — some trailing jobs missing; dedup picks them up later)"
        print(msg)

    elif argv[0] == "dept":
        if len(argv) < 4:
            print("Usage: extract-cache dept <slug> <dept_id> <input_file>")
            sys.exit(1)
        slug, dept_id, input_path = argv[1], argv[2], argv[3]
        jobs, err = gh.extract_dept_from_raw(Path(input_path).read_text(errors="replace"))
        out_dir = cdir / "boards"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{slug}-dept-{dept_id}.json"
        out_path.write_text(json.dumps({"jobs": jobs}, indent=2))
        msg = f"Wrote {len(jobs)} job(s) to {out_path}"
        if err:
            msg += f" ({err})"
        print(msg)

    else:  # posting
        job_id, input_path = argv[1], argv[2]
        obj, status = gh.extract_posting_from_raw(Path(input_path).read_text(errors="replace"))
        out_dir = cdir / "postings"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{job_id}.json"
        if status == "ok":
            out_path.write_text(json.dumps(obj, indent=2))
            print(f"Wrote posting {job_id} to {out_path}")
        else:
            out_path.write_text(json.dumps({"_status": status}))
            print(f"Posting {job_id}: {status} response - wrote placeholder to {out_path}")


def cmd_list_departments(argv):
    """python3 daily_pipeline.py list-departments <slug> <input_file>

    Parse a web_fetch result of GET /v1/boards/{slug}/departments and print
    a table of department IDs, names, and job counts. Use this to find the
    right IDs to put in the Departments column of company-boards.xlsx.

    Workflow (run once per company, not every day):
      1. web_fetch https://boards-api.greenhouse.io/v1/boards/{slug}/departments
      2. Save result to a file
      3. python3 daily_pipeline.py list-departments {slug} <result_file>
      4. Copy the IDs for Engineering/IT departments into company-boards.xlsx

    For more than one company, prefer `list-departments-batch` instead -
    this single-board command has no file output and is easy to lose
    track of across many invocations."""
    if len(argv) < 2:
        print(cmd_list_departments.__doc__)
        sys.exit(1)
    slug, input_path = argv[0], argv[1]
    text = Path(input_path).read_text(errors="replace")
    depts, err = gh.extract_deptlist_from_raw(text)
    if err:
        print(f"WARNING: {err}")
    if not depts:
        print("No departments found in response.")
        sys.exit(0)
    print(f"\nDepartments for board '{slug}' ({len(depts)} total):\n")
    print(f"{'ID':>10}  {'Jobs':>5}  Name")
    print("-" * 60)
    for d in sorted(depts, key=lambda x: x.get("name", "")):
        print(f"{d.get('id', ''):>10}  {len(d.get('jobs', [])):>5}  {d.get('name', '')}")
    print(f"\nAdd desired IDs (comma-separated) to the 'Departments' column in company-boards.xlsx.")
    print(f"Leave blank to pull all jobs from this board (no department filter).")


def cmd_derive_departments_batch(argv):
    """python3 daily_pipeline.py derive-departments-batch <slug1> <slug2> ...
       python3 daily_pipeline.py derive-departments-batch <slug1>=<file1> ...  (fallback mode)

    Alternative to list-departments-batch when /departments keeps getting
    truncated even via direct fetch (e.g. outbound HTTPS is blocked in this
    environment too, so it's falling back to a web_fetch-sourced cache file
    that was ALREADY truncated, and re-running the same command against the
    same file can't produce more complete data — there's nothing new to
    extract). This command sidesteps /departments and instead fetches
    /jobs?content=true and derives department id/name/job-count directly
    from the jobs returned.

    IMPORTANT: this needs content=true, not content=false. Greenhouse only
    includes the `departments` array on a job object when content=true -
    with content=false (the daily pipeline's normal setting, used to avoid
    the description-field weight on every run) departments is omitted
    entirely. An earlier version of this command used content=false and
    got 0 departments back for every company despite fetching successfully -
    confirmed by a real run with no truncation and no errors, just an empty
    result. If you're running this from your own machine (recommended -
    see discover_departments.py) there's no fetch-size cap to worry about,
    so content=true costs nothing extra here.

    One real limitation vs. /departments: a department with zero current
    postings won't appear (there's no job carrying that department to
    derive it from). For finding which departments to filter TO going
    forward, this is rarely a problem - an empty department isn't useful
    for the Departments column anyway.

    Same transport strategy as list-departments-batch: bare slugs try
    direct Python fetch first (no size cap), falling back to a cache file
    only if direct fails.
    """
    if len(argv) < 1:
        print(cmd_derive_departments_batch.__doc__)
        sys.exit(1)

    lines = ["DEPARTMENT DISCOVERY REPORT (derived from /jobs, not /departments)", "=" * 60, ""]
    any_warning = False
    needs_manual_fetch = []

    for arg in argv:
        if "=" in arg:
            slug, input_path = arg.split("=", 1)
            if not Path(input_path).exists():
                lines.append(f"\n--- {slug} ---")
                lines.append(f"ERROR: file not found: {input_path}")
                any_warning = True
                continue
            text = Path(input_path).read_text(errors="replace")
            jobs, truncated = gh.extract_jobs_from_raw(text)
            err = "truncated - some trailing jobs missing, department list may be incomplete" if truncated else None
            source = "cache file"
        else:
            slug = arg
            jobs, err, source = gh.fetch_jobs_with_content(slug)

        lines.append(f"\n--- {slug} --- (source: {source}, {len(jobs)} job(s) retrieved)")
        if err:
            lines.append(f"WARNING: {err}")
            any_warning = True
            if source == "none":
                needs_manual_fetch.append(slug)
        depts = gh.derive_departments_from_jobs(jobs)
        if not depts:
            lines.append("No departments found in the retrieved jobs.")
            continue
        lines.append(f"{len(depts)} department(s) seen in these jobs:")
        lines.append(f"{'ID':>10}  {'Jobs':>5}  Name")
        lines.append("-" * 60)
        for d in depts:
            lines.append(f"{d['id']:>10}  {d['count']:>5}  {d['name']}")

    lines.append("")
    lines.append("=" * 60)
    if any_warning:
        lines.append("Some boards had issues - see WARNING lines above per company.")
        lines.append("Job counts shown ARE exact for the jobs actually retrieved (no '?'")
        lines.append("marking needed), but a truncated/capped fetch means some jobs - and")
        lines.append("possibly entire departments with no other postings - may be missing.")
    if needs_manual_fetch:
        lines.append("")
        lines.append("These companies need a manual web_fetch (direct Python fetch failed")
        lines.append("AND no cache file was found) - fetch each, save to a file, then re-run")
        lines.append("this command with slug=file for just these:")
        for slug in needs_manual_fetch:
            lines.append(f"  {slug}: web_fetch {gh.jobs_with_content_url(slug)}")
    lines.append("")
    lines.append("Add desired IDs (comma-separated) to the 'Departments' column in")
    lines.append("company-boards.xlsx for each company. Leave blank to keep pulling")
    lines.append("all jobs from that board (no department filter).")

    report = "\n".join(lines)
    out_path = DATA_DIR / "department_discovery_report.txt"
    out_path.write_text(report)
    print(report)
    print(f"\n(Also written to {out_path} in case this gets cut off in chat.)")


def cmd_list_departments_batch(argv):
    """python3 daily_pipeline.py list-departments-batch <slug1> <slug2> ...
       python3 daily_pipeline.py list-departments-batch <slug1>=<file1> ...  (fallback mode)

    Batch version of list-departments: discovers departments for multiple
    companies in ONE call, aggregates into ONE report (printed in full AND
    written to department_discovery_report.txt).

    TRANSPORT: for each bare slug (no "="), this tries gh.fetch_departments()
    FIRST - a direct Python HTTP call with no response-size cap at all. This
    is the actual fix for the truncation problem that affects Cowork's
    web_fetch (~94KB cap) on this endpoint: a direct call has no such cap,
    so there's nothing to truncate in the first place. Only falls back to a
    pre-fetched cache file (or an explicit slug=file argument) if the direct
    call fails for that company - and even then, the cache file may itself
    be a truncated web_fetch dump, which is when the recovery parsing in
    extract_deptlist_from_raw kicks in as a last resort.

    Do NOT default to having Cowork web_fetch every board first - that
    reintroduces the exact truncation problem this command exists to avoid.
    Pass bare slugs and let this command attempt direct fetch on its own;
    only fall back to web_fetch for companies this command itself reports
    as failed.

    SCOPE: only pass slugs for companies that genuinely need discovery -
    i.e. NOT already BLACKLIST or WATCH in company-tracker.xlsx. The
    daily-job-scrub-pipeline.md README task is responsible for filtering
    the slug list down to this before calling this command.
    """
    if len(argv) < 1:
        print(cmd_list_departments_batch.__doc__)
        sys.exit(1)

    lines = ["DEPARTMENT DISCOVERY REPORT", "=" * 60, ""]
    any_warning = False
    needs_manual_fetch = []

    for arg in argv:
        if "=" in arg:
            # Fallback mode: read a pre-fetched file directly, skip the
            # direct-fetch attempt entirely (caller already has the data).
            slug, input_path = arg.split("=", 1)
            if not Path(input_path).exists():
                lines.append(f"\n--- {slug} ---")
                lines.append(f"ERROR: file not found: {input_path}")
                any_warning = True
                continue
            text = Path(input_path).read_text(errors="replace")
            depts, err = gh.extract_deptlist_from_raw(text)
            source = "cache file"
        else:
            # Default mode: try direct fetch first.
            slug = arg
            depts, err, source = gh.fetch_departments(slug)

        lines.append(f"\n--- {slug} --- (source: {source})")
        if err:
            lines.append(f"WARNING: {err}")
            any_warning = True
            if source == "none":
                needs_manual_fetch.append(slug)
        if not depts:
            lines.append("No departments found.")
            continue
        lines.append(f"{len(depts)} department(s):")
        lines.append(f"{'ID':>10}  {'Jobs':>5}  Name")
        lines.append("-" * 60)
        for d in sorted(depts, key=lambda x: x.get("name", "")):
            jobs_str = str(len(d.get("jobs", [])))
            if err:  # job counts are unreliable when truncation recovery kicked in
                jobs_str += "?"
            lines.append(f"{d.get('id', ''):>10}  {jobs_str:>5}  {d.get('name', '')}")

    lines.append("")
    lines.append("=" * 60)
    if any_warning:
        lines.append("Some boards had issues - see WARNING lines above per company.")
        lines.append("Truncated-but-recovered responses: job counts marked '?' are")
        lines.append("unreliable (reflect only what fit before any cut), but department")
        lines.append("IDs and names are still accurate.")
    if needs_manual_fetch:
        lines.append("")
        lines.append("These companies need a manual web_fetch (direct Python fetch failed")
        lines.append("AND no cache file was found) - fetch each, save to a file, then re-run")
        lines.append("this command with slug=file for just these:")
        for slug in needs_manual_fetch:
            lines.append(f"  {slug}: web_fetch {gh.departments_url(slug)}")
    lines.append("")
    lines.append("Add desired IDs (comma-separated) to the 'Departments' column in")
    lines.append("company-boards.xlsx for each company. Leave blank to keep pulling")
    lines.append("all jobs from that board (no department filter).")

    report = "\n".join(lines)
    out_path = DATA_DIR / "department_discovery_report.txt"
    out_path.write_text(report)
    print(report)
    print(f"\n(Also written to {out_path} in case this gets cut off in chat.)")





# ---------------------------------------------------------------------------
# Issues summary (printed at the end of each phase)
# ---------------------------------------------------------------------------

def print_issues_summary(c1=None, c2=None):
    """Either 'No script issues.' or a detailed bulleted list - this is the
    format Cowork should also follow for issues on its own side (web_fetch
    failures, file permission errors, etc.) when reporting back."""
    issues = []
    if c1:
        for e in c1.get("board_errors", []):
            issues.append(f"Board fetch/cache: {e}")
        if c1.get("capped"):
            issues.append(
                f"Task 1 hit max_postings_per_run ({c1.get('fetched')} fetched) - "
                f"some boards/postings weren't processed this run; they'll be picked up next run."
            )
    if c2:
        if c2.get("revisit_refetch_failed"):
            issues.append(
                f"{c2['revisit_refetch_failed']} pending-revisit posting(s) couldn't be "
                f"refreshed after their company resolved - marked 'check manually' in seen-postings."
            )

    print()
    if not issues:
        print("No script issues.")
    else:
        print(f"Script issues ({len(issues)}):")
        for i in issues:
            print(f"  - {i}")


# ---------------------------------------------------------------------------
# Review & Apply: the one sheet B actually needs to look at
# ---------------------------------------------------------------------------

_REVIEW_FIT_RANK = {"Strong Match": 0, "Solid Contender": 1, "Long Shot": 2, "Hard No": 3}
_REVIEW_ROLE_RANK = {"PURSUE": 0, "CHECK": 1}
_REVIEW_PAY_RANK = {"OK": 0, "Below Target": 1, "Unusually High": 2, "Not Disclosed": 3, "Too Low": 4}


def _parse_requirements_ratio(s):
    """"7/9" -> 0.778. Returns None if unparseable (e.g. "N/A")."""
    m = re.match(r"\s*(\d+)\s*/\s*(\d+)\s*", str(s or ""))
    if not m:
        return None
    num, den = int(m.group(1)), int(m.group(2))
    return (num / den) if den else None


def _count_csv_terms(s):
    s = str(s or "").strip()
    if not s or s.lower() in ("none", "-", "n/a"):
        return 0
    return len([x for x in s.split(",") if x.strip()])


def _review_rank_key(shortlist_d, fit_d):
    """Lower sorts first = better. Tiers: Company Fit, then Role Match, then
    (for assessed rows) Requirements Met ratio / Pay Flag / ATS critical-gap
    count. Unassessed rows sit just behind assessed ones in the same
    Company Fit + Role Match tier - they're not penalized on quality (which
    we don't know yet), only ordered after rows we've actually verified."""
    fit_rank = _REVIEW_FIT_RANK.get(shortlist_d.get("Company Fit"), 9)
    role_rank = _REVIEW_ROLE_RANK.get(shortlist_d.get("Role Match"), 9)
    if fit_d:
        ratio = _parse_requirements_ratio(fit_d.get("Requirements Met (Required)"))
        ratio_rank = -(ratio if ratio is not None else 0.5)
        pay_rank = _REVIEW_PAY_RANK.get(fit_d.get("Pay Flag"), 9)
        ats_rank = _count_csv_terms(fit_d.get("ATS Keywords - Missing (Critical)"))
        assessed_rank = 0
    else:
        ratio_rank, pay_rank, ats_rank, assessed_rank = -0.5, 9, 9, 1
    return (fit_rank, role_rank, assessed_rank, ratio_rank, pay_rank, ats_rank)


def build_review_sheet(cfg, wb_listings):
    """Rebuild "Shortlist" (the ranked list B actually reviews): All Listings rows with Status not in
    (Applied, Skipped), joined with Fit Assessment data, ranked best-first,
    capped at cfg["review_list_size"]. Call sync_review_to_shortlist() first
    so this run's edits are reflected in the Status filter. Returns the list
    of rows written (for the final run summary)."""
    ws_short = wb_listings["All Listings"]
    ws_fit = wb_listings["Fit Assessment"]

    fit_by_jid, fit_by_url = {}, {}
    for _, fd in rows_as_dicts(ws_fit):
        if fd.get("Job ID"):
            fit_by_jid[str(fd["Job ID"])] = fd
        if fd.get("Posting URL"):
            fit_by_url[fd["Posting URL"]] = fd

    candidates = []
    for _, sd in rows_as_dicts(ws_short):
        status = str(sd.get("Status") or "").strip().lower()
        if status in ("applied", "skipped"):
            continue
        jid = sd.get("Job ID")
        fd = fit_by_jid.get(str(jid)) if jid else None
        if fd is None:
            fd = fit_by_url.get(sd.get("URL"))
        candidates.append((sd, fd))

    candidates.sort(key=lambda pair: _review_rank_key(pair[0], pair[1]))

    # Cap how many postings per company can appear on the Shortlist at once.
    # Without this, a company with many open roles (Asana, Figma) can fill
    # most or all of the list, crowding out other companies' postings even
    # when they rank similarly well - the Shortlist is meant to show a
    # spread of options worth pursuing, not an exhaustive per-company dump.
    per_company_cap = cfg["max_shortlist_per_company"]
    company_counts: dict = {}
    capped_candidates = []
    for sd, fd in candidates:
        co = normalize_company_name(sd.get("Company Name") or "").casefold()
        if company_counts.get(co, 0) >= per_company_cap:
            continue
        company_counts[co] = company_counts.get(co, 0) + 1
        capped_candidates.append((sd, fd))
    candidates = capped_candidates[: cfg["review_list_size"]]

    ws_review = wb_listings["Shortlist"]
    for r in range(ws_review.max_row, 1, -1):
        ws_review.delete_rows(r)

    rows_written = []
    for i, (sd, fd) in enumerate(candidates, start=1):
        row = {
            "Rank": i,
            "Company Name": sd.get("Company Name"),
            "Job Title": sd.get("Job Title"),
            "Department": sd.get("Department") or "Not Disclosed",
            "Company Fit": sd.get("Company Fit"),
            "Role Match": sd.get("Role Match"),
            "Requirements Met (Required)": fd.get("Requirements Met (Required)") if fd else "Not yet assessed",
            "Pay Range": (fd.get("Pay Range") if fd else None) or sd.get("Pay Range") or "Not Disclosed",
            "Pay Flag": fd.get("Pay Flag") if fd else "Not yet assessed",
            "Key Gaps": fd.get("Key Gaps") if fd else "",
            "Location": sd.get("Location"),
            "Location Type": sd.get("Location Type"),
            "URL": sd.get("URL"),
            "Job ID": sd.get("Job ID"),
            "Date Found": sd.get("Date Found"),
            "Posted Date": sd.get("Posted Date") or "Not Disclosed",
            "Job Description": sd.get("Job Description") or "",
            "Status": "",
            "Skip Reason": "",
        }
        append_dict_row(ws_review, row)
        rows_written.append(row)
    return rows_written


def vetting_cost_report(wb_tracker, vetted_companies=None):
    """Returns a list of {"company": str, "chars": int, "words": int} for
    rows freshly vetted this run (Date Checked == TODAY, or explicitly
    listed in vetted_companies if provided).

    This is NOT a token count - daily_pipeline.py has no access to actual
    token usage, since that's tracked by the Claude API / Cowork's own
    infrastructure on the model side of the conversation, not in any file
    or interface this script can see. Character/word count of what
    actually got WRITTEN to company-tracker.xlsx is a proxy: it can't tell
    you the exact token cost of a company-vetting-subagent.md run (which
    also includes the web searches/fetches themselves, not just the final
    write), but it's a real, measurable signal for "which companies are
    expensive to vet" and "did the field-trimming actually reduce output."
    Words are roughly chars/4.5 (rough English average, not exact - good
    enough to eyeball relative cost between companies and runs, not
    precise enough to budget against)."""
    ws_tracker = wb_tracker["Company Tracker"]
    target_companies = None
    if vetted_companies is not None:
        target_companies = {normalize_company_name(c).casefold() for c in vetted_companies}

    report = []
    for _, d in rows_as_dicts(ws_tracker):
        name = d.get("Company Name")
        if target_companies is not None:
            if normalize_company_name(name).casefold() not in target_companies:
                continue
        else:
            date_checked = _parse_date(d.get("Date Checked"))
            if date_checked != TODAY:
                continue
        chars = sum(len(str(v)) for v in d.values() if v)
        report.append({"company": name, "chars": chars, "words": round(chars / 4.5)})
    return report


def companies_needing_vetting_despite_dept_setup(wb_boards, wb_tracker):
    """Returns a list of company names that HAVE department IDs configured
    in company-boards.xlsx but have NO row at all in company-tracker.xlsx -
    i.e. department discovery happened (or department IDs were set up some
    other way) before the company was ever actually vetted PURSUE/WATCH/
    BLACKLIST. This is the mirror case of companies_needing_department_discovery:
    that one catches "vetted but not filtered", this one catches "filtered
    but never vetted" - both are signs the two steps happened out of order,
    but this one is the more important gap to close, since a board can sit
    here indefinitely contributing postings without ever having been
    through the vetting subagent at all."""
    ws_tracker = wb_tracker["Company Tracker"]
    tracked_companies = {
        normalize_company_name(d.get("Company Name")).casefold()
        for _, d in rows_as_dicts(ws_tracker)
    }
    ws_boards = wb_boards["Company Boards"]
    needs_vetting = []
    for _, d in rows_as_dicts(ws_boards):
        name = d.get("Company Name")
        key = normalize_company_name(name).casefold()
        if key not in tracked_companies and str(d.get("Departments") or "").strip():
            needs_vetting.append(name)
    return needs_vetting


def companies_needing_department_discovery(wb_boards, wb_tracker):
    """Returns a list of company names that are PURSUE in company-tracker.xlsx
    but have a blank Departments column in company-boards.xlsx - i.e. newly
    vetted companies that will keep pulling their entire unfiltered job list
    every run until someone runs department discovery for them.

    This is the actual signal for "should I run discover_departments.py
    again" - not a fixed schedule, just whenever this list is non-empty."""
    ws_tracker = wb_tracker["Company Tracker"]
    pursue_companies = {
        normalize_company_name(d.get("Company Name")).casefold()
        for _, d in rows_as_dicts(ws_tracker)
        if str(d.get("Status")).strip().upper() == "PURSUE"
    }
    ws_boards = wb_boards["Company Boards"]
    needs_discovery = []
    for _, d in rows_as_dicts(ws_boards):
        name = d.get("Company Name")
        key = normalize_company_name(name).casefold()
        if key in pursue_companies and not str(d.get("Departments") or "").strip():
            needs_discovery.append(name)
    return needs_discovery


def print_run_summary(phase1_report, phase2_report, review_rows, needs_dept_discovery=None, needs_vetting_despite_setup=None, wb_tracker=None):
    """The final human-facing summary. Format:
    - Initial retrieval counts
    - New companies vetted → check company-tracker.xlsx
    - Unclassified titles → check job-title-filters.xlsx
    - Unrecognized locations → check location-filters.xlsx
    - Final shortlist count + top 3 preview
    """
    t1 = phase1_report.get("task1", {})
    t3 = phase1_report.get("task3", {})

    print(f"\nInitial Retrieval: {t1.get('fetched', 0)} job listing(s) "
          f"from {t1.get('boards_attempted', 0)} company board(s) "
          f"→ {t1.get('already_seen', 0)} already seen, "
          f"{t1.get('location_excluded', 0)} wrong location, "
          f"{t1.get('title_skip', 0)} title-skipped "
          f"→ {t1.get('survivors', 0)} new listing(s) this run.")

    vetted = t3.get("sent_to_vetting", [])
    if vetted:
        print(f"\n{len(vetted)} new compan{'y' if len(vetted) == 1 else 'ies'} vetted — "
              f"review verdicts in company-tracker.xlsx:")
        for v in vetted:
            print(f"  - {v['company']}")
        if wb_tracker is not None:
            cost_report = vetting_cost_report(wb_tracker, vetted_companies=[v["company"] for v in vetted])
            if cost_report:
                total_words = sum(r["words"] for r in cost_report)
                print(f"  Output size this run (proxy for vetting cost, NOT actual tokens used — "
                      f"see vetting_cost_report's docstring for why an exact count isn't possible): "
                      f"~{total_words} words across {len(cost_report)} compan{'y' if len(cost_report) == 1 else 'ies'}")
                for r in sorted(cost_report, key=lambda x: -x["words"]):
                    print(f"    - {r['company']}: ~{r['words']} words ({r['chars']} chars)")
    else:
        print("\nNo new companies vetted this run.")

    if needs_dept_discovery:
        print(f"\n{len(needs_dept_discovery)} PURSUE compan{'y' if len(needs_dept_discovery) == 1 else 'ies'} "
              f"still pulling their FULL job list (no department filter set up yet):")
        for name in needs_dept_discovery:
            print(f"  - {name}")
        print(f"  Run discover_departments.py for {'this one' if len(needs_dept_discovery) == 1 else 'these'} "
              f"when convenient, then add the picks to company-boards.xlsx \"Departments\" column. "
              f"Not urgent — the unfiltered pull still works, just less efficiently.")

    if needs_vetting_despite_setup:
        print(f"\n{len(needs_vetting_despite_setup)} compan{'y' if len(needs_vetting_despite_setup) == 1 else 'ies'} "
              f"{'has' if len(needs_vetting_despite_setup) == 1 else 'have'} department IDs configured "
              f"in company-boards.xlsx but {'was' if len(needs_vetting_despite_setup) == 1 else 'were'} "
              f"never actually vetted (no row in company-tracker.xlsx at all):")
        for name in needs_vetting_despite_setup:
            print(f"  - {name}")
        print(f"  Run company-vetting-subagent.md for {'this one' if len(needs_vetting_despite_setup) == 1 else 'these'} "
              f"— department filtering was set up before vetting happened, so right now "
              f"{'it is' if len(needs_vetting_despite_setup) == 1 else 'they are'} sitting in the lower-priority "
              f"tier alongside WATCH companies despite the filtering work already done.")

    unclassified = t1.get("unclassified_titles", 0)
    if unclassified:
        print(f"\n{unclassified} unclassified job title(s) logged — "
              f"please review job-title-filters.xlsx → \"Unclassified Titles\" "
              f"and add each to the Title Keywords sheet as SKIP/PURSUE/CHECK.")

    unrecognized_locs = t1.get("unrecognized_locations", 0)
    if unrecognized_locs:
        print(f"\n{unrecognized_locs} unrecognized location(s) logged — "
              f"please review location-filters.xlsx → \"Unrecognized Locations\" "
              f"and add any new regions to the Location Keywords sheet.")

    total = (phase2_report or {}).get("fit_assessment", {}).get("needs_assessment", len(review_rows))
    assessed = sum(1 for r in review_rows if r.get("Requirements Met (Required)") != "Not yet assessed")
    print(f"\n{len(review_rows)} listing(s) in job-listings.xlsx → \"Shortlist\" "
          f"({assessed} with full fit assessment, {len(review_rows) - assessed} not yet assessed).")
    if review_rows:
        print("Top 3:")
        for row in review_rows[:3]:
            req = row.get("Requirements Met (Required)", "")
            req_str = f", Req: {req}" if req and req != "Not yet assessed" else ""
            pay = row.get("Pay Flag", "")
            pay_str = f", Pay: {pay}" if pay and pay != "Not yet assessed" else ""
            print(f"  {row['Rank']}. {row['Company Name']} — {row['Job Title']} "
                  f"({row['Company Fit']} / {row['Role Match']}{req_str}{pay_str})")

    print(f"\nOpen job-listings.xlsx → \"Shortlist\" to review. "
          f"Set Status = Applied or Skipped (+ optional Skip Reason) for each. "
          f"Changes sync automatically on the next run.")


# ---------------------------------------------------------------------------
# Phase 1: Tasks 0, 1, 2, Task 3 triage
# ---------------------------------------------------------------------------

def phase1(cfg):
    print("=== Phase 1: Tasks 0, 1, 2, Task 3 triage ===\n")

    wb_listings = load_workbook(DATA_DIR / "job-listings.xlsx")
    wb_seen = load_workbook(DATA_DIR / "seen-postings.xlsx")
    wb_titles = load_workbook(DATA_DIR / "job-title-filters.xlsx")
    wb_loc = load_workbook(DATA_DIR / "location-filters.xlsx")
    wb_boards = load_workbook(DATA_DIR / "company-boards.xlsx")
    wb_sources = load_workbook(DATA_DIR / "job-sources.xlsx")
    wb_tracker = load_workbook(DATA_DIR / "company-tracker.xlsx")

    # One-time / idempotent: normalize any datetime or ISO-string date cells
    # to plain date objects so Excel shows clean dates, not timestamps.
    _migrated = 0
    for wb, fname in [
        (wb_listings, "job-listings.xlsx"), (wb_seen, "seen-postings.xlsx"),
        (wb_titles, "job-title-filters.xlsx"), (wb_loc, "location-filters.xlsx"),
        (wb_tracker, "company-tracker.xlsx"),
    ]:
        _migrated += migrate_dates_in_workbook(wb, DATE_COLUMNS.get(fname, {}))
    if _migrated:
        print(f"Migrated {_migrated} date cell(s) to plain date format (one-time cleanup).")

    review_synced = sync_review_to_shortlist(wb_listings)
    if review_synced:
        print(f"Synced {review_synced} Status/Skip Reason edit(s) from \"Shortlist\" to All Listings.")

    updated = task0_sync_applied_status(wb_listings, wb_seen)
    print(f"Task 0: {updated} seen-postings row(s) synced to Applied?=Yes")

    survivors, c = task1_pull_dedup_filter(cfg, wb_boards, wb_sources, wb_seen, wb_titles, wb_loc, wb_tracker)
    print(
        f"Task 1: {c['fetched']} fetched ({c['boards_attempted']} boards"
        + (f", {len(c['board_errors'])} errors" if c['board_errors'] else "")
        + f", capped={c['capped']})"
        + f" -> {c['no_url']} to Needs Link, {c['already_seen']} already seen, {c['new']} new"
        + f" -> {c['location_excluded']} location-excluded, {c['stale_excluded']} stale-excluded, {c['title_skip']} title-SKIP"
        + f" -> {c['survivors']} survivors"
    )
    src_counts = {}
    for company, src in c.get("fetch_sources", {}).items():
        src_counts[src] = src_counts.get(src, 0) + 1
    if src_counts:
        breakdown = ", ".join(f"{n} {src}" for src, n in sorted(src_counts.items()))
        print(f"  fetch source: {breakdown}")
        if src_counts.get("direct", 0) > 0:
            print(f"  ({src_counts['direct']} board(s) fetched directly via Python - "
                  f"no Cowork web_fetch needed, no truncation risk)")
    for e in c["board_errors"]:
        print(f"  board error - {e}")
    if "rotation_offset_this_run" in c:
        print(f"  PURSUE-board rotation: started at PURSUE board #{c['rotation_offset_this_run'] + 1} this run, "
              f"next run starts at #{c['rotation_offset_next_run'] + 1}"
              + (" (cap hit before reaching every board - it'll get there over a few runs)" if c["capped"] else ""))

    groups = task2_group_by_company(survivors, wb_tracker)
    carried_over = sum(1 for _, cnt in groups if cnt == 0)
    print(f"Task 2: {len(groups)} unique companies"
          + (f" (incl. {carried_over} carried-over PENDING from a prior run)" if carried_over else ""))

    handoff, resolved, pending_written = task3_triage(cfg, groups, wb_tracker)
    print(f"Task 3 triage: {len(groups)} companies -> {len(resolved)} resolved instantly, "
          f"{len(handoff)} sent to vetting subagent, {len(pending_written)} written to PENDING (capped)")

    wb_seen.save(DATA_DIR / "seen-postings.xlsx")
    wb_titles.save(DATA_DIR / "job-title-filters.xlsx")
    wb_loc.save(DATA_DIR / "location-filters.xlsx")
    wb_tracker.save(DATA_DIR / "company-tracker.xlsx")
    wb_listings.save(DATA_DIR / "job-listings.xlsx")

    (DATA_DIR / "survivors.json").write_text(json.dumps(survivors, indent=2, default=str))
    (DATA_DIR / "handoff_companies.json").write_text(json.dumps(handoff, indent=2))
    (DATA_DIR / "phase1_report.json").write_text(json.dumps({
        "date": TODAY.isoformat(),
        "task0_applied_synced": updated,
        "task1": c,
        "task2_company_groups": len(groups),
        "task3": {
            "resolved_instantly": resolved,
            "sent_to_vetting": handoff,
            "written_to_pending": pending_written,
        },
    }, indent=2, default=str))

    print()
    if handoff:
        print(f"Companies needing the vetting subagent this run ({len(handoff)}):")
        for h in handoff:
            print(f"  - {h['company']} ({h['listing_count']} listing(s))")
        print("\nNext: run company-vetting-subagent.md for each company above, "
              "writing its verdict to company-tracker.xlsx. Then run:")
        print("  python3 daily_pipeline.py phase2")
    else:
        print("No companies need vetting this run. Next: python3 daily_pipeline.py phase2")

    print_issues_summary(c1=c)


# ---------------------------------------------------------------------------
# Phase 2: Task 4 + Fit Assessment handoff
# ---------------------------------------------------------------------------

def phase2(cfg):
    print("=== Phase 2: Task 4 ===\n")

    survivors_path = DATA_DIR / "survivors.json"
    if not survivors_path.exists():
        print("ERROR: survivors.json not found. Run 'python3 daily_pipeline.py phase1' first.")
        sys.exit(1)
    survivors = json.loads(survivors_path.read_text())
    # Posted Date round-trips through JSON as a string (date isn't JSON-native) -
    # restore it to a real date object so All Listings gets a proper date cell,
    # not a text string. "Not Disclosed" and unparseable values pass through as-is.
    for s in survivors:
        pd = s.get("Posted Date")
        if isinstance(pd, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", pd):
            try:
                s["Posted Date"] = date.fromisoformat(pd)
            except ValueError:
                pass

    wb_tracker = load_workbook(DATA_DIR / "company-tracker.xlsx")
    wb_listings = load_workbook(DATA_DIR / "job-listings.xlsx")
    wb_seen = load_workbook(DATA_DIR / "seen-postings.xlsx")
    wb_boards = load_workbook(DATA_DIR / "company-boards.xlsx")

    c = task4_filter_shortlist(cfg, survivors, wb_tracker, wb_listings, wb_seen, wb_boards)
    print(
        f"Task 4: {c['candidates']} candidate(s) -> {c['added']} added to Shortlist  "
        f"(discarded: {c['discarded_blacklist']} blacklist, {c['discarded_watch']} watch, "
        f"{c['pending_deferred']} pending-deferred, {c['fit_floor_excluded']} below fit floor)"
    )
    for company, total, n in c["app_limit_capped"]:
        print(f"  application limit: {company} {total} matched, capped to {n}")
    if c["revisit_added"] or c["revisit_cleared"] or c["revisit_refetch_failed"]:
        print(
            f"  pending-revisit rescan: {c['revisit_added']} added, {c['revisit_cleared']} cleared, "
            f"{c['revisit_refetch_failed']} refetch failed"
        )

    wb_listings.save(DATA_DIR / "job-listings.xlsx")
    wb_seen.save(DATA_DIR / "seen-postings.xlsx")

    # --- Fit Assessment handoff ---
    ws_short = wb_listings["All Listings"]
    ws_fit = wb_listings["Fit Assessment"]
    assessed_jids = {str(d.get("Job ID")) for _, d in rows_as_dicts(ws_fit) if d.get("Job ID")}
    assessed_urls = {d.get("Posting URL") for _, d in rows_as_dicts(ws_fit) if d.get("Posting URL")}

    needs_assessment = []
    for _, d in rows_as_dicts(ws_short):
        jid, url = d.get("Job ID"), d.get("URL")
        if (jid and str(jid) in assessed_jids) or (url and url in assessed_urls):
            continue
        needs_assessment.append(d)

    fit_rank = {"Strong Match": 0, "Solid Contender": 1, "Long Shot": 2, "Hard No": 3}

    def _staleness_penalty(d):
        """Tiebreaker among postings that already passed the hard
        max_posting_age_days cutoff in task1 (so this never sees anything
        older than that cutoff allows) - prefers fresher postings within
        that window when multiple postings are competing for the same
        fit-assessment slot. 0 if Posted Date is missing/unparseable."""
        pd = d.get("Posted Date")
        if isinstance(pd, str):
            try:
                pd = date.fromisoformat(pd[:10])
            except ValueError:
                return 0
        if not isinstance(pd, date):
            return 0
        age_days = (TODAY - pd).days
        if age_days > 2:
            return 2
        if age_days > 1:
            return 1
        return 0

    def _title_quality_penalty(title):
        """Posting-level signal beyond the SKIP list: titles with senior/lead
        qualifiers correlate with poor fit even when not hard-skipped (e.g. a
        title that matched PURSUE on "Software Engineer" but is actually
        "Senior Staff Software Engineer, Tech Lead"). Mild penalty, not a skip -
        the title filters already removed the clear non-fits."""
        t = (title or "").lower()
        penalty = 0
        for kw in ["staff", "principal", "lead ", "tech lead", "architect", "fellow"]:
            if kw in t:
                penalty += 1
        return penalty

    needs_assessment.sort(key=lambda d: (
        fit_rank.get(d.get("Company Fit"), 9),
        0 if d.get("Role Match") == "PURSUE" else 1,
        _title_quality_penalty(d.get("Job Title")),
        _staleness_penalty(d),
        # Within the same tier, sort company name alphabetically so we
        # round-robin across companies rather than exhausting one company first.
        # The per-company cap below enforces the actual spread.
        normalize_company_name(d.get("Company Name") or "").casefold(),
    ))

    # Spread slots across companies: at most ~1/4 of the cap per company, so
    # a single large board (e.g. Anthropic with 139 listings) can't
    # monopolize all the fit-assessment slots for a run.
    cap = cfg["max_fit_assessments_per_run"]
    per_company_cap = max(2, cap // 4)  # e.g. cap=15 -> max 3 per company
    company_counts: dict = {}
    handoff_fit = []
    for d in needs_assessment:
        if len(handoff_fit) >= cap:
            break
        co = normalize_company_name(d.get("Company Name") or "").casefold()
        if company_counts.get(co, 0) >= per_company_cap:
            continue
        company_counts[co] = company_counts.get(co, 0) + 1
        handoff_fit.append(d)

    # If we didn't fill the cap (too few companies have eligible postings),
    # raise the per-company cap incrementally rather than ignoring it outright -
    # this still favors spread over depth, just relaxes the limit one step at a
    # time only as far as needed to fill the cap.
    relax = per_company_cap
    while len(handoff_fit) < cap and relax < cap:
        relax += 1
        company_counts = {}
        handoff_fit = []
        for d in needs_assessment:
            if len(handoff_fit) >= cap:
                break
            co = normalize_company_name(d.get("Company Name") or "").casefold()
            if company_counts.get(co, 0) >= relax:
                continue
            company_counts[co] = company_counts.get(co, 0) + 1
            handoff_fit.append(d)

    print(f"\nTask 5 handoff: {len(needs_assessment)} Shortlist row(s) need assessment, "
          f"{len(handoff_fit)} selected this run (cap={cap}, max {relax}/company"
          f"{' - relaxed from ' + str(per_company_cap) if relax != per_company_cap else ''})")

    (DATA_DIR / "handoff_fit_assessments.json").write_text(json.dumps([
        {"Job Title": d["Job Title"], "Company Name": d["Company Name"],
         "Job ID": d["Job ID"], "Posting URL": d["URL"],
         "Company Fit": d.get("Company Fit"), "Role Match": d.get("Role Match")}
        for d in handoff_fit
    ], indent=2))
    (DATA_DIR / "phase2_report.json").write_text(json.dumps({
        "date": TODAY.isoformat(),
        "task4": c,
        "fit_assessment": {
            "needs_assessment": len(needs_assessment),
            "selected_this_run": len(handoff_fit),
            "cap": cap,
        },
    }, indent=2, default=str))

    if handoff_fit:
        print("\nNext: run job-fit-assessment-subagent.md for each row in handoff_fit_assessments.json, "
              "writing results to job-listings.xlsx (\"Fit Assessment\").")
    else:
        print("\nNo new Shortlist rows need assessment this run.")

    print_issues_summary(c2=c)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1] == "extract-cache":
        cmd_extract_cache(sys.argv[2:])
        sys.exit(0)

    if len(sys.argv) >= 2 and sys.argv[1] == "extract-cache-batch":
        cmd_extract_cache_batch(sys.argv[2:])
        sys.exit(0)

    if len(sys.argv) >= 2 and sys.argv[1] == "list-departments":
        cmd_list_departments(sys.argv[2:])
        sys.exit(0)

    if len(sys.argv) >= 2 and sys.argv[1] == "list-departments-batch":
        cmd_list_departments_batch(sys.argv[2:])
        sys.exit(0)

    if len(sys.argv) >= 2 and sys.argv[1] == "derive-departments-batch":
        cmd_derive_departments_batch(sys.argv[2:])
        sys.exit(0)

    if len(sys.argv) >= 2 and sys.argv[1] == "write-description":
        # Usage: python3 daily_pipeline.py write-description <job_id>
        # Reads the HTML content straight from the posting cache file
        # ({cache_dir}/postings/{job_id}.json, written by `extract-cache posting`)
        # — no need to pass HTML as a command-line argument, which is fragile
        # for multi-KB content (shell quoting/length limits).
        if len(sys.argv) < 3:
            print("Usage: python3 daily_pipeline.py write-description <job_id> [--cache-dir DIR]")
            sys.exit(1)
        job_id = sys.argv[2]
        cdir_override = sys.argv[sys.argv.index("--cache-dir") + 1] if "--cache-dir" in sys.argv else None
        posting_path = gh.cache_dir(cdir_override) / "postings" / f"{job_id}.json"
        if not posting_path.exists():
            print(f"ERROR: No cached posting at {posting_path}. Fetch it first: "
                  f"web_fetch the job detail URL (?content=true), then "
                  f"python3 daily_pipeline.py extract-cache posting {job_id} <result_file>")
            sys.exit(1)
        posting_data = json.loads(posting_path.read_text())
        if posting_data.get("_status") in ("empty", "truncated"):
            print(f"Posting {job_id}: cached response was {posting_data['_status']} - no description to write.")
            sys.exit(0)
        html_content = posting_data.get("content", "")
        wb_listings = load_workbook(DATA_DIR / "job-listings.xlsx")
        found = write_job_description(wb_listings, job_id, None, html_content)
        if found:
            wb_listings.save(DATA_DIR / "job-listings.xlsx")
            chars = len(strip_html(html_content))
            print(f"Wrote {chars} char description for Job ID {job_id}.")
        else:
            print(f"WARNING: Job ID {job_id} not found in All Listings - description not written.")
        sys.exit(0)

    if len(sys.argv) >= 2 and sys.argv[1] == "build-review":
        cfg = load_config()
        wb_listings = load_workbook(DATA_DIR / "job-listings.xlsx")

        synced = sync_review_to_shortlist(wb_listings)
        if synced:
            print(f"Synced {synced} Status/Skip Reason edit(s) from \"Shortlist\" to All Listings.")

        review_rows = build_review_sheet(cfg, wb_listings)
        wb_listings.save(DATA_DIR / "job-listings.xlsx")
        print(f"Wrote {len(review_rows)} row(s) to \"Shortlist\" (cap={cfg['review_list_size']}).")

        p1_path, p2_path = DATA_DIR / "phase1_report.json", DATA_DIR / "phase2_report.json"
        if p1_path.exists() and p2_path.exists():
            phase1_report = json.loads(p1_path.read_text())
            phase2_report = json.loads(p2_path.read_text())
            wb_boards = load_workbook(DATA_DIR / "company-boards.xlsx")
            wb_tracker = load_workbook(DATA_DIR / "company-tracker.xlsx")
            needs_dept_discovery = companies_needing_department_discovery(wb_boards, wb_tracker)
            needs_vetting_despite_setup = companies_needing_vetting_despite_dept_setup(wb_boards, wb_tracker)
            print_run_summary(phase1_report, phase2_report, review_rows, needs_dept_discovery, needs_vetting_despite_setup, wb_tracker)
        else:
            print("\n(phase1_report.json / phase2_report.json not found - run phase1 and phase2 "
                  "first for the full run summary.)")
        sys.exit(0)

    if len(sys.argv) != 2 or sys.argv[1] not in ("phase1", "phase2"):
        print(__doc__)
        sys.exit(1)

    cfg = load_config()
    if sys.argv[1] == "phase1":
        phase1(cfg)
    else:
        phase2(cfg)
